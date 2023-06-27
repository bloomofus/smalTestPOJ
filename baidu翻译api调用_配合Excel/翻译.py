import random
import time
from hashlib import md5

import openpyxl
import requests


def translate_text(text, from_lang, to_lang):
    # Set your own appid/appkey.
    appid = '20230625001724138'
    appkey = 'pwopu8lAz4mimvXcO2HX'

    # For list of language codes, please refer to `https://api.fanyi.baidu.com/doc/21`
    from_lang = from_lang
    to_lang = to_lang

    endpoint = 'http://api.fanyi.baidu.com'
    path = '/api/trans/vip/translate'
    url = endpoint + path

    query = text

    # Generate salt and sign
    def make_md5(s, encoding='utf-8'):
        return md5(s.encode(encoding)).hexdigest()

    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)

    # Build request
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}

    # Send request
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()

    # get response
    result = result['trans_result']
    tempStr = ''
    for line in result:
        tempStr += line['dst']
    result = tempStr
    # 延时1s，因为免费api的速率限制
    time.sleep(1)
    return result


def trans(inputPath, inputLie, outputLie, startRow=1, outputPath=None, from_lang='en', to_lang='zh'):
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(inputPath)
    worksheet = workbook.active
    # 列和数字对应字典
    dic = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7}  # 修改字典对应的列序号

    # 确保输入列和输出列存在于字典中
    if not {inputLie, outputLie}.issubset(dic):
        raise Exception("输入错误")

    inputCol = dic[inputLie]  # 列号，从1开始
    outputCol = dic[outputLie]

    # 确保目标列存在
    max_column = worksheet.max_column  # 获取表格的最大列号,列号从1开始
    if outputCol > max_column:
        # 如果目标列不存在，则添加一列，在第startRow行添加临时文本
        worksheet.cell(row=startRow, column=outputCol, value='tempTxt')
    # 逐行翻译并填写到 B 列
    for row in worksheet.iter_rows(min_row=startRow):
        text_to_translate = row[inputCol - 1].value
        if text_to_translate == None:
            continue  # 考虑到存在空格的情况
        translated_text = translate_text(text_to_translate, from_lang, to_lang)
        row[outputCol - 1].value = translated_text

    if outputPath != None:
        try:
            # 保存翻译结果到新文件
            workbook.save(outputPath)
        except:
            raise Exception("输出路径错误")
    else:
        # 保存翻译至原来文件
        workbook.save(inputPath)


trans(inputPath='./xlsx/temp.xlsx', inputLie='A', outputLie='B',
      startRow=1, outputPath='./xlsx/temp2.xlsx', from_lang='en', to_lang='zh')
